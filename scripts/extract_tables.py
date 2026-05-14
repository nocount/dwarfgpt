"""
Step 2 of the port plan: turn the Excel workbook into the canonical JSON the
Python module will read at runtime.

Outputs (all UTF-8 NFC-normalized JSON):

  data/dictionary.json
      [{"english_with_tag": str, "english_clean": str, "tag": str | null,
        "khuzdul_raw": str, "khuzdul_clean": str, "row": int}]
      ~215K entries from CONSTRUCT!A2:C214948.
      english_clean / tag are derived from english_with_tag by splitting on the
      first '[' (same idea as the workbook's column C formula).
      khuzdul_clean is khuzdul_raw with the workbook's three markup strips
      applied (_  -  (ul)) so it's the same string the phonetic pipeline sees.

  data/phonemes.json
      [{"latin": str, "am_code": str, "ae_code": str,
        "moria_font_char": str, "erebor_font_char": str,
        "rune_name": str | null, "phonetic_category": str | null,
        "ipa": str, "ipa_west": str | null, "ipa_central": str | null,
        "sounds_like": str | null, "pronunciation_tip": str | null}]
      The 53 letter rows of Phonetics and Cirth!A3:P55, plus the full-stop row 56.

  data/digraphs.json
      [{"first_am": str, "second_am": str, "result_am": str,
        "first_latin": str, "second_latin": str, "result_latin": str}]
      The 16 digraph collapse rules parsed out of the Converter!CK6 formula.

  data/phonological_rules.json
      {
        "end_shwa":   [{"from": str, "to": str, "id": str | null, "row": int}, ...],
        "mid_shwa":   [...],
        "caret":      [...],
        "gemination": [...]
      }
      The rule tables from the four column groups of Shwa-Caret-Gemination.

  data/raw/
      raw_<sheet>.values.tsv      cell-value dumps (data_only=True)
      raw_<sheet>.formulas.tsv    formula-string dumps (data_only=False)
      For provenance only; the Python runtime reads only the JSON above.

Throughput: the 215K dictionary scan is the only slow part (~30 s). Everything
else completes in a few seconds.
"""

from __future__ import annotations

import io
import json
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "sentence_maker_original_v1.xlsm"
DATA = ROOT / "data"
RAW = DATA / "raw"


def nfc(s: Any) -> Any:
    if isinstance(s, str):
        return unicodedata.normalize("NFC", s)
    return s


def log(msg: str) -> None:
    print(f"[extract] {msg}", flush=True)


# ---------------------------------------------------------------------------
# Dictionary
# ---------------------------------------------------------------------------

# The workbook's column-C cleaning formula is
#   =SUBSTITUTE(LEFT(A2, FIND("[", A2, 1)), "[", "")
# i.e. take everything up to and including the FIRST "[", then drop the "[".
# For entries with no "[" the formula errors and the displayed value is the
# unchanged string. We match that behavior, then also split the bracketed
# annotation into (grammatical tag, radical marker) when the entry has two
# bracket groups like `... [<tag>] [<radical>]`.
_BRACKET_GROUP_RE = re.compile(r"\[([^\[\]]*)\]")


def split_english(english_with_tag: str) -> tuple[str, str | None, str | None]:
    """Return (clean, tag, radical). Tag and radical are None when absent.

    `clean`   — text before the first '['; matches the workbook's column C.
    `tag`     — content of the first '[...]' group (grammatical annotation).
    `radical` — content of the LAST '[...]' group if there are two or more,
                typically the all-caps consonantal root marker like 'KhZD'.
                Returned only when at least two bracket groups exist (so a
                single-tag entry returns radical=None).
    """
    if not isinstance(english_with_tag, str):
        return ("", None, None)
    s = english_with_tag
    first_bracket = s.find("[")
    if first_bracket < 0:
        return (nfc(s.strip()), None, None)
    clean = nfc(s[:first_bracket].strip())
    groups = [m.group(1).strip() for m in _BRACKET_GROUP_RE.finditer(s)]
    tag = nfc(groups[0]) if groups else None
    radical = nfc(groups[-1]) if len(groups) >= 2 else None
    return (clean, tag, radical)


# The workbook applies three markup strips to the Khuzdul before romanization:
#   SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(AL6, "_", ""), "-", ""), "(ul)", "")
# We replicate exactly that order.
def strip_khuzdul_markup(khuzdul: str) -> str:
    if not isinstance(khuzdul, str):
        return ""
    s = khuzdul.replace("_", "").replace("-", "").replace("(ul)", "")
    return nfc(s)


def extract_dictionary(wb_v) -> list[dict[str, Any]]:
    log("Extracting dictionary from CONSTRUCT (read_only pass for speed)...")
    # Re-load in read_only mode for a faster scan of column A/B.
    t0 = time.time()
    wb_ro = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb_ro["CONSTRUCT"]
    entries: list[dict[str, Any]] = []
    row_idx = 0
    blank_run = 0
    BLANK_THRESHOLD = 2000  # generous: the dictionary has no internal gaps this large
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2, values_only=True):
        row_idx += 1
        actual_row = row_idx + 1  # 1-based, accounting for header
        a, b = row[0], row[1]
        if (a is None or (isinstance(a, str) and not a.strip())) and (
            b is None or (isinstance(b, str) and not b.strip())
        ):
            blank_run += 1
            if blank_run >= BLANK_THRESHOLD:
                break
            continue
        blank_run = 0
        english_with_tag = nfc(a) if isinstance(a, str) else ""
        khuzdul_raw = nfc(b) if isinstance(b, str) else ""
        clean, tag, radical = split_english(english_with_tag)
        entries.append(
            {
                "row": actual_row,
                "english_with_tag": english_with_tag,
                "english_clean": clean,
                "tag": tag,
                "radical": radical,
                "khuzdul_raw": khuzdul_raw,
                "khuzdul_clean": strip_khuzdul_markup(khuzdul_raw),
            }
        )
        if row_idx % 20000 == 0:
            log(f"  ... scanned {row_idx} rows, kept {len(entries)} ({time.time() - t0:.1f}s)")
    wb_ro.close()
    log(f"Dictionary: {len(entries)} entries ({time.time() - t0:.1f}s)")
    return entries


# ---------------------------------------------------------------------------
# Phonemes (Phonetics and Cirth sheet)
# ---------------------------------------------------------------------------


def extract_phonemes(wb_v) -> list[dict[str, Any]]:
    log("Extracting phonemes from Phonetics and Cirth sheet ...")
    ws = wb_v["Phonetics and Cirth"]
    # Columns of interest (1-indexed):
    #   A: AM Code      B: AE Code      C: letter to type (Latin)
    #   D: Moria glyph  E: letter to type (duplicate)
    #   F: Erebor glyph G: letter to type (duplicate)
    #   H: rune name    I: AM Code (dup)  J: AE Code (dup)
    #   K: phonetic category   L,M,N: IPA / phonetic symbol variants
    #   O: sounds like   P: tips to pronounce
    rows: list[dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        am_code = ws.cell(row=r, column=1).value
        if am_code is None or not str(am_code).strip():
            continue
        rune_name = str(ws.cell(row=r, column=8).value or "").strip()
        # The source cells wrap the rune name in literal double quotes,
        # e.g. "'om" or "siginan" (long A). Strip a single matched pair.
        if rune_name.startswith('"') and rune_name.find('"', 1) > 0:
            close = rune_name.find('"', 1)
            inner = rune_name[1:close]
            trailing = rune_name[close + 1 :].strip()
            rune_name = inner if not trailing else f"{inner} {trailing}"
        rows.append(
            {
                "row": r,
                "am_code": nfc(str(am_code).strip()),
                "ae_code": nfc(str(ws.cell(row=r, column=2).value or "").strip()) or None,
                "latin": nfc(str(ws.cell(row=r, column=3).value or "").strip()),
                "moria_font_char": nfc(str(ws.cell(row=r, column=4).value or "")),
                "erebor_font_char": nfc(str(ws.cell(row=r, column=6).value or "")),
                "rune_name": nfc(rune_name) or None,
                "rune_name_raw": nfc(str(ws.cell(row=r, column=8).value or "").strip()) or None,
                "phonetic_category": nfc(str(ws.cell(row=r, column=11).value or "").strip()) or None,
                "ipa": nfc(str(ws.cell(row=r, column=12).value or "").strip()),
                "ipa_west": nfc(str(ws.cell(row=r, column=13).value or "").strip()) or None,
                "ipa_central": nfc(str(ws.cell(row=r, column=14).value or "").strip()) or None,
                "sounds_like": nfc(str(ws.cell(row=r, column=15).value or "").strip()) or None,
                "pronunciation_tip": nfc(str(ws.cell(row=r, column=16).value or "").strip()) or None,
            }
        )
    log(f"Phonemes: {len(rows)} entries")
    return rows


# ---------------------------------------------------------------------------
# Digraphs (parsed from Converter!CK6 formula)
# ---------------------------------------------------------------------------

# CK6 looks like:
#   IF(OR(CK6=...digraph_codes...), "",
#     IF(AND(BQ6="AM038", BR6="AM014"), "AM039",
#       IF(AND(BQ6="AM038", BR6="AM036"), "AM040",
#         ... 16 levels deep ...
#         BQ6)))
# The pattern we care about is the chain of AND/quoted-codes triples.
_DIGRAPH_RE = re.compile(
    r'AND\(B[A-Z]+\d+="(AM\d+)"\s*,\s*B[A-Z]+\d+="(AM\d+)"\)\s*,\s*"(AM\d+)"',
    re.IGNORECASE,
)


def extract_digraphs(wb_f, phonemes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    log("Extracting digraph rules from Converter!CK6 ...")
    ws = wb_f["Converter"]
    formula = ws["CK6"].value
    if not isinstance(formula, str):
        raise RuntimeError("Converter!CK6 is not a formula; cannot parse digraphs.")
    code_to_latin = {p["am_code"]: p["latin"] for p in phonemes}
    seen: set[tuple[str, str, str]] = set()
    rules: list[dict[str, Any]] = []
    for m in _DIGRAPH_RE.finditer(formula):
        first, second, result = m.group(1), m.group(2), m.group(3)
        key = (first, second, result)
        if key in seen:
            continue
        seen.add(key)
        rules.append(
            {
                "first_am": first,
                "second_am": second,
                "result_am": result,
                "first_latin": code_to_latin.get(first, ""),
                "second_latin": code_to_latin.get(second, ""),
                "result_latin": code_to_latin.get(result, ""),
            }
        )
    log(f"Digraphs: {len(rules)} rules")
    return rules


# ---------------------------------------------------------------------------
# Phonological rules (Shwa-Caret-Gemination)
# ---------------------------------------------------------------------------


def extract_rule_family(
    ws,
    src_col: int,
    dst_col: int,
    id_col: int | None,
    *,
    family: str,
    max_scan: int = 200_000,
) -> list[dict[str, Any]]:
    """Walk two columns of the sheet collecting (from, to) pairs."""
    rules: list[dict[str, Any]] = []
    blank_run = 0
    for r in range(2, max_scan + 1):
        src = ws.cell(row=r, column=src_col).value
        dst = ws.cell(row=r, column=dst_col).value
        if src is None and dst is None:
            blank_run += 1
            if blank_run >= 50:
                break
            continue
        blank_run = 0
        if src is None or dst is None:
            # A one-sided row is a data oddity; skip but log
            continue
        rid = ws.cell(row=r, column=id_col).value if id_col else None
        rules.append(
            {
                "row": r,
                "id": nfc(str(rid).strip()) if rid is not None else None,
                "from": nfc(str(src)),
                "to": nfc(str(dst)),
                "family": family,
            }
        )
    return rules


def extract_phonological_rules(wb_v) -> dict[str, list[dict[str, Any]]]:
    log("Extracting phonological rules from Shwa-Caret-Gemination ...")
    ws = wb_v["Shwa-Caret-Gemination"]
    # Column groups identified during inspection:
    #   END SHWA:   A (from), B (to), C (rule id W*)
    #   MID SHWA:   J (from), K (to)            [no rule-id column]
    #   Caret:      CC (from), CD (to)          [CE = formula cell that applies the chain,
    #                                            with "skip if English contains 'Plural'" guard]
    #   Gemination: CH (from), CI (to)          [CJ = formula cell that applies the chain
    #                                            and is what CONSTRUCT row 4 reads]
    end_shwa = extract_rule_family(ws, 1, 2, 3, family="end_shwa")
    log(f"  end_shwa: {len(end_shwa)}")
    mid_shwa = extract_rule_family(ws, 10, 11, None, family="mid_shwa")
    log(f"  mid_shwa: {len(mid_shwa)}")
    caret = extract_rule_family(ws, 81, 82, None, family="caret")
    log(f"  caret: {len(caret)}")
    gemination = extract_rule_family(ws, 86, 87, None, family="gemination")
    log(f"  gemination: {len(gemination)}")
    return {
        "end_shwa": end_shwa,
        "mid_shwa": mid_shwa,
        "caret": caret,
        "gemination": gemination,
    }


# ---------------------------------------------------------------------------
# Raw dumps (provenance)
# ---------------------------------------------------------------------------


def dump_raw_sheets(wb_v, wb_f, sheet_names: list[str], *, cap_rows: int = 1000) -> None:
    """For non-huge sheets only. Writes tab-separated dumps to data/raw/."""
    RAW.mkdir(parents=True, exist_ok=True)
    for name in sheet_names:
        log(f"  raw dump: {name}")
        for kind, wb in (("values", wb_v), ("formulas", wb_f)):
            ws = wb[name]
            safe = name.replace(" ", "_").replace("/", "_")
            out = RAW / f"raw_{safe}.{kind}.tsv"
            with out.open("w", encoding="utf-8") as f:
                end_row = min(ws.max_row, cap_rows)
                for r in range(1, end_row + 1):
                    cells = []
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v is None:
                            cells.append("")
                        else:
                            s = str(v).replace("\t", " ").replace("\n", " ")
                            cells.append(s)
                    f.write("\t".join(cells))
                    f.write("\n")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def write_json(path: Path, payload: Any, *, indent: int | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=indent)
        f.write("\n")
    size = path.stat().st_size
    log(f"  wrote {path.relative_to(ROOT)} ({size:,} bytes)")


def main() -> int:
    log(f"Loading workbook (values) from {WORKBOOK.name} ...")
    t0 = time.time()
    wb_v = load_workbook(WORKBOOK, data_only=True, keep_vba=True)
    log(f"  loaded values in {time.time() - t0:.1f}s")
    log("Loading workbook (formulas) ...")
    t0 = time.time()
    wb_f = load_workbook(WORKBOOK, data_only=False, keep_vba=True)
    log(f"  loaded formulas in {time.time() - t0:.1f}s")

    DATA.mkdir(exist_ok=True)
    RAW.mkdir(exist_ok=True)

    # 1. Dictionary (uses its own read_only load for speed)
    dictionary = extract_dictionary(wb_v)
    # Sort by row (preserves the workbook's ordering, which is roughly alphabetical)
    write_json(DATA / "dictionary.json", dictionary)

    # 2. Phonemes
    phonemes = extract_phonemes(wb_v)
    write_json(DATA / "phonemes.json", phonemes, indent=2)

    # 3. Digraphs (needs phonemes to back-fill latin)
    digraphs = extract_digraphs(wb_f, phonemes)
    write_json(DATA / "digraphs.json", digraphs, indent=2)

    # 4. Phonological rules
    rules = extract_phonological_rules(wb_v)
    write_json(DATA / "phonological_rules.json", rules)

    # 5. Raw provenance dumps for the small sheets
    dump_raw_sheets(
        wb_v,
        wb_f,
        ["Phonetics and Cirth", "CONCAT", "FINAL"],
        cap_rows=1000,
    )

    # Summary
    log("---- summary ----")
    log(f"dictionary entries        : {len(dictionary):,}")
    log(f"phoneme rows              : {len(phonemes)}")
    log(f"digraph rules             : {len(digraphs)}")
    log(f"end-shwa rules            : {len(rules['end_shwa'])}")
    log(f"mid-shwa rules            : {len(rules['mid_shwa'])}")
    log(f"caret rules               : {len(rules['caret'])}")
    log(f"gemination rules          : {len(rules['gemination'])}")
    log("done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
Second-pass inspection: zoom into the rectangles that matter.

- CONSTRUCT: columns A..G of the first 8 rows (header zone), the H..AK live-input
  zone (rows 1..10), and the column G labels down rows 1..12.
- Phonetics and Cirth: dump the full sheet.
- Shwa-Caret-Gemination: dump the column headers and a sample of the rewrite table.
- CONCAT: dump formulas verbatim.
- FINAL: dump everything (it's tiny).
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

WORKBOOK = Path(__file__).resolve().parent.parent / "sentence_maker_original_v1.xlsm"


def show_rect(ws, fws, row_lo: int, row_hi: int, col_lo: int, col_hi: int, *, label: str) -> None:
    print(f"### {label}  — rows {row_lo}..{row_hi}, cols {col_lo}..{col_hi}\n")
    print("```")
    # column header row (Excel letters)
    from openpyxl.utils import get_column_letter
    header = "  r |" + "|".join(f" {get_column_letter(c):>18} " for c in range(col_lo, col_hi + 1))
    print(header)
    print("-" * len(header))
    for r in range(row_lo, row_hi + 1):
        cells = []
        for c in range(col_lo, col_hi + 1):
            v = ws.cell(row=r, column=c).value
            cells.append("" if v is None else str(v).replace("\n", " ")[:18])
        print(f"{r:>3} |" + "|".join(f" {x:>18} " for x in cells))
    print("```\n")

    # formulas in this rectangle
    formulas: list[tuple[str, str, object]] = []
    for r in range(row_lo, row_hi + 1):
        for c in range(col_lo, col_hi + 1):
            f = fws.cell(row=r, column=c).value
            if isinstance(f, str) and f.startswith("="):
                v = ws.cell(row=r, column=c).value
                formulas.append((fws.cell(row=r, column=c).coordinate, f, v))
    if formulas:
        print("**Formulas in this rectangle:**\n")
        print("```")
        for coord, f, v in formulas[:40]:
            vs = "" if v is None else str(v).replace("\n", " ")[:24]
            print(f"{coord:>6}  {f[:100]:<100}  =>  {vs}")
        if len(formulas) > 40:
            print(f"... ({len(formulas) - 40} more formulas not shown)")
        print("```\n")


def main() -> int:
    wb_v = load_workbook(WORKBOOK, data_only=True, keep_vba=True)
    wb_f = load_workbook(WORKBOOK, data_only=False, keep_vba=True)

    print("# Deep inspection\n")

    # ---- CONSTRUCT ----
    print("## CONSTRUCT\n")
    ws = wb_v["CONSTRUCT"]
    fws = wb_f["CONSTRUCT"]

    show_rect(ws, fws, 1, 12, 1, 7, label="CONSTRUCT — dictionary + label column (A..G, top 12 rows)")
    show_rect(ws, fws, 1, 12, 8, 37, label="CONSTRUCT — H..AK lookup zone (rows 1..12)")
    show_rect(ws, fws, 1, 12, 38, 43, label="CONSTRUCT — AL..AQ (rows 1..12)")

    # last-row sanity
    print(f"- CONSTRUCT max_row = {ws.max_row}  (sheet probably has lots of blank trailing rows)\n")

    # How many real dictionary rows are there? Count non-empty in column A starting at row 2.
    # We stop counting at the first run of, say, 200 consecutive blanks to keep this fast.
    print("- Counting dictionary rows by scanning column A...")
    real_count = 0
    last_real_row = 0
    blank_run = 0
    BLANK_THRESHOLD = 500
    # use iter_rows for speed
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None or (isinstance(v, str) and not v.strip()):
            blank_run += 1
            if blank_run >= BLANK_THRESHOLD:
                break
        else:
            real_count += 1
            last_real_row = 2 + real_count + (blank_run if last_real_row == 0 else 0)  # rough only
            blank_run = 0
    print(f"  - non-blank rows in column A (approximate): {real_count}")
    print()

    # ---- Phonetics and Cirth ----
    print("## Phonetics and Cirth — full sheet\n")
    ws = wb_v["Phonetics and Cirth"]
    fws = wb_f["Phonetics and Cirth"]
    show_rect(ws, fws, 1, ws.max_row, 1, ws.max_column, label="Phonetics and Cirth — all")

    # ---- Shwa-Caret-Gemination header + sample ----
    print("## Shwa-Caret-Gemination — first 20 rows × first 16 cols\n")
    ws = wb_v["Shwa-Caret-Gemination"]
    fws = wb_f["Shwa-Caret-Gemination"]
    show_rect(ws, fws, 1, 20, 1, 16, label="Shwa-Caret-Gemination — top-left corner")

    # ---- CONCAT full ----
    print("## CONCAT — full sheet (visible final-assembly)\n")
    ws = wb_v["CONCAT"]
    fws = wb_f["CONCAT"]
    show_rect(ws, fws, 1, ws.max_row, 1, ws.max_column, label="CONCAT — all")

    # Print the full formulas (not truncated) from CONCAT — they are the heart of the assembly.
    print("**Full formulas in CONCAT:**\n")
    print("```")
    for row in fws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                print(f"{cell.coordinate}: {cell.value}")
                print()
    print("```\n")

    # ---- FINAL full ----
    print("## FINAL — full sheet\n")
    ws = wb_v["FINAL"]
    fws = wb_f["FINAL"]
    show_rect(ws, fws, 1, ws.max_row, 1, ws.max_column, label="FINAL — all")

    # ---- Converter sample ----
    print("## Converter — rows 1..20, cols AL..BD\n")
    ws = wb_v["Converter"]
    fws = wb_f["Converter"]
    # AL = 38, BD = 56
    show_rect(ws, fws, 1, 20, 38, 56, label="Converter — top of input zone")
    print("**Full Converter formulas, row 6:**\n")
    print("```")
    for c in range(1, ws.max_column + 1):
        f = fws.cell(row=6, column=c).value
        if isinstance(f, str) and f.startswith("="):
            from openpyxl.utils import get_column_letter
            print(f"{get_column_letter(c)}6: {f}")
    print("```\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
